"""Shared bulk import/export helpers (Phase 2).

Read and write tabular data uniformly as CSV or XLSX so every module's bulk
endpoints behave the same. Importers get a list of ``{lower-cased-header: str}``
rows; exporters/templates get back ``(bytes, media_type, ext)``.
"""

from __future__ import annotations

import csv
import io
from collections.abc import Callable
from datetime import date, datetime
from typing import Any

from openpyxl import Workbook, load_workbook

from app.core.exceptions import BadRequest
from app.schemas.common_bulk import ImportError as ImportErrorRow
from app.schemas.common_bulk import ImportResult

# --- cell value parsers (tolerant; reused by every importer) ---
_DATE_FORMATS = (
    "%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d %b %Y", "%d %B %Y",
)
_TRUE_VALUES = {"1", "true", "yes", "y", "t", "paid", "done", "x"}
_BLANKS = {"", "-", "--", "n/a", "na", "none"}


def parse_date(value: str) -> date | None:
    text = (value or "").strip()
    if not text:
        return None
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"Unrecognized date '{value}' (use YYYY-MM-DD)")


def parse_bool(value: str) -> bool:
    return (value or "").strip().lower() in _TRUE_VALUES


def parse_number(value: str) -> float | None:
    text = (value or "").strip().replace(",", "").replace("$", "").replace("₹", "")
    if text.lower() in _BLANKS:
        return None
    try:
        return float(text)
    except ValueError as exc:
        raise ValueError(f"Invalid number '{value}'") from exc


def parse_int(value: str) -> int | None:
    num = parse_number(value)
    return int(num) if num is not None else None

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CSV_MEDIA = "text/csv"


def _norm(value: Any) -> str:
    """Coerce a cell to a trimmed string (integers stay integer-looking)."""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_xlsx(filename: str, content: bytes) -> bool:
    name = (filename or "").lower()
    if name.endswith((".xlsx", ".xlsm")):
        return True
    if name.endswith(".csv"):
        return False
    return content[:2] == b"PK"  # xlsx is a zip


def parse_table(filename: str, content: bytes) -> list[dict[str, str]]:
    """Parse CSV or XLSX bytes into header-keyed (lower-cased) string rows."""
    if _looks_xlsx(filename, content):
        return _parse_xlsx(content)
    return _parse_csv(content)


def _parse_csv(content: bytes) -> list[dict[str, str]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError as exc:
        raise BadRequest("File must be UTF-8 encoded CSV (or upload an .xlsx)") from exc
    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        return []
    return [
        {(k or "").strip().lower(): _norm(v) for k, v in raw.items()} for raw in reader
    ]


def _parse_xlsx(content: bytes) -> list[dict[str, str]]:
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 - any openpyxl failure => bad file
        raise BadRequest("Could not read the Excel file") from exc
    ws = wb.active
    if ws is None:
        wb.close()
        return []
    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter, None)
    if header is None:
        wb.close()
        return []
    headers = [_norm(h).lower() for h in header]
    out: list[dict[str, str]] = []
    for row in rows_iter:
        if row is None or all(c is None for c in row):
            continue
        record: dict[str, str] = {}
        for i, key in enumerate(headers):
            if key:
                record[key] = _norm(row[i]) if i < len(row) else ""
        out.append(record)
    wb.close()
    return out


def write_table(
    columns: list[str], rows: list[list[Any]], fmt: str
) -> tuple[bytes, str, str]:
    """Serialize rows to CSV or XLSX. Returns (content, media_type, extension)."""
    if fmt == "xlsx":
        wb = Workbook()
        ws = wb.active
        ws.append(columns)
        for row in rows:
            ws.append(["" if c is None else c for c in row])
        # Light styling: bold header.
        for cell in ws[1]:
            cell.font = cell.font.copy(bold=True)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue(), XLSX_MEDIA, "xlsx"
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(columns)
    for row in rows:
        writer.writerow(["" if c is None else c for c in row])
    return buf.getvalue().encode("utf-8-sig"), CSV_MEDIA, "csv"


def template(
    columns: list[str], example: list[Any] | None, fmt: str
) -> tuple[bytes, str, str]:
    """A header-only (or header + one example row) starter file."""
    return write_table(columns, [example] if example else [], fmt)


def normalize_format(fmt: str | None) -> str:
    value = (fmt or "csv").lower()
    return "xlsx" if value in ("xlsx", "excel", "xls") else "csv"


def run_row_imports(
    db: Any, rows: list[dict[str, str]], handler: Callable[[dict[str, str]], bool]
) -> ImportResult:
    """Apply ``handler`` to each row inside its own savepoint, collecting errors.

    ``handler(row)`` returns True when it created a record, False when it updated
    one, and may raise to flag a bad row (isolated; other rows still import).
    """
    created = 0
    updated = 0
    errors: list[ImportErrorRow] = []
    for i, row in enumerate(rows, start=2):  # row 1 is the header
        try:
            with db.begin_nested():
                was_created = handler(row)
            if was_created:
                created += 1
            else:
                updated += 1
        except Exception as exc:  # noqa: BLE001 - per-row isolation
            errors.append(ImportErrorRow(row=i, message=str(exc)))
    return ImportResult(created=created, updated=updated, errors=errors)
