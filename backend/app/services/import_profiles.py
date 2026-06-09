"""Reusable import profiles (Phase 2 import engine).

A *profile* knows how to turn a specific workbook format into canonical rows for
one entity, how to validate + de-duplicate them, and how to apply them to the DB.
Profiles are reusable: the same profile handles every future upload of that
format. Today: two Projects profiles (the clean template + the team's real
"Projects & Assignee" master tab). Payments profiles land next.
"""

from __future__ import annotations

import io
import uuid
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.exceptions import BadRequest
from app.models.client import Client
from app.models.lookups import Country, Niche
from app.models.payment import Payment, PaymentStatusHistory
from app.models.project import Project
from app.models.user import User
from app.services.bulk import parse_date, parse_number, parse_table

PAYMENT_STATUSES = {"pending", "negotiation", "paid", "free", "cancelled", "rejected"}

PROJECT_STATUSES = {"active", "completed", "hold", "cancelled"}

# Loose aliases so messy real-world country/niche cells still resolve.
COUNTRY_ALIASES = {
    "usa": "us", "u.s.a": "us", "u.s.": "us", "united states of america": "us",
    "uk": "gb", "u.k.": "gb", "england": "gb", "britain": "gb",
    "uae": "ae", "u.a.e": "ae", "emirates": "ae",
}


def _excel_serial_to_date(serial: float) -> date:
    # Excel's day 0 is 1899-12-30 (accounts for the 1900 leap-year bug).
    return date(1899, 12, 30) + timedelta(days=int(serial))


def _coerce_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, int | float):
        return _excel_serial_to_date(value)
    return parse_date(str(value))  # may raise ValueError


@dataclass
class Issue:
    level: str  # "error" blocks the row; "warning" still imports
    message: str


@dataclass
class ExtractedRow:
    row_number: int
    raw: dict[str, Any]
    canonical: dict[str, Any]
    source: str | None = None  # e.g. originating sheet/tab name


@dataclass
class ApplyOutcome:
    action: str  # created | updated
    entity_id: uuid.UUID
    old_snapshot: dict[str, Any] | None = None


def _read_named_sheet(content: bytes, *candidates: str) -> tuple[list[str], list[tuple[int, list[Any]]]]:
    """Read the first matching sheet (by case-insensitive name) with NATIVE cell
    values. Returns (headers, [(excel_row_number, values), ...])."""
    try:
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        raise BadRequest("Could not read the Excel file") from exc
    wanted = {c.strip().lower() for c in candidates}
    target = None
    for name in wb.sheetnames:
        if name.strip().lower() in wanted:
            target = name
            break
    if target is None:
        wb.close()
        raise BadRequest(
            f"Could not find a '{candidates[0]}' sheet (tabs: {', '.join(wb.sheetnames[:12])}…)"
        )
    ws = wb[target]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], []
    headers = [str(h).strip() if h is not None else "" for h in rows[0]]
    data = [(i, list(r)) for i, r in enumerate(rows[1:], start=2)]
    return headers, data


@dataclass
class ResolveContext:
    niches: dict[str, Niche] = field(default_factory=dict)
    countries: dict[str, Country] = field(default_factory=dict)
    users: dict[str, User] = field(default_factory=dict)
    projects_by_name: dict[str, Project] = field(default_factory=dict)


class ProjectProfileBase:
    """Shared validate/dedupe/apply for any profile that yields Project rows."""

    entity_type = "project"
    on_duplicate = "update"  # re-importing the same project updates it

    def exists(self, key: str, ctx: ResolveContext) -> bool:
        return key in ctx.projects_by_name

    def build_context(self, db: Session, company_id: uuid.UUID) -> ResolveContext:
        ctx = ResolveContext()
        for n in db.query(Niche).all():
            ctx.niches[n.name.strip().lower()] = n
        for c in db.query(Country).all():
            ctx.countries[c.iso_code.lower()] = c
            ctx.countries[c.name.lower()] = c
        for u in db.query(User).filter(User.company_id == company_id).all():
            ctx.users[u.email.lower()] = u
            ctx.users[u.full_name.strip().lower()] = u  # also match by name
        for p in db.query(Project).filter(Project.company_id == company_id).all():
            ctx.projects_by_name[p.name.strip().lower()] = p
        return ctx

    def dedupe_key(self, canonical: dict[str, Any]) -> str:
        return (canonical.get("name") or "").strip().lower()

    def validate(self, canonical: dict[str, Any], ctx: ResolveContext) -> list[Issue]:
        issues: list[Issue] = []
        name = (canonical.get("name") or "").strip()
        if not name:
            issues.append(Issue("error", "Project name is required"))
            return issues
        for field_name, label in (("main_niche", "Main niche"), ("project_niche", "Project niche")):
            val = (canonical.get(field_name) or "").strip()
            if val and self._match_niche(val, ctx) is None:
                issues.append(Issue("warning", f"{label} '{val}' not matched — left blank"))
        country = (canonical.get("target_country") or "").strip()
        if country and self._match_country(country, ctx) is None:
            issues.append(Issue("warning", f"Country '{country}' not matched — left blank"))
        for field_name, label in (("assignee", "Assignee"), ("team_lead", "Team lead")):
            val = (canonical.get(field_name) or "").strip()
            if val and ctx.users.get(val.lower()) is None:
                issues.append(Issue("warning", f"{label} '{val}' not matched to a user — left blank"))
        if not (canonical.get("assignee") or "").strip():
            issues.append(Issue("warning", "No assignee specified"))
        status = (canonical.get("status") or "").strip().lower()
        if status and status not in PROJECT_STATUSES:
            issues.append(Issue("warning", f"Unknown status '{status}' — defaulted to active"))
        try:
            self._number(canonical.get("monthly_budget"))
        except ValueError:
            issues.append(Issue("warning", "Monthly budget isn't a plain number — left as 0"))
        try:
            links = self._number(canonical.get("target_links"))
            if links is not None and links > 100000:
                issues.append(
                    Issue("warning", f"Target links '{links:.0f}' looks like a date, not a count — ignored")
                )
        except ValueError:
            issues.append(Issue("warning", "Target links isn't a number — ignored"))
        due = canonical.get("due_date")
        if isinstance(due, date) and not (2000 <= due.year <= 2100):
            issues.append(Issue("warning", f"Due date {due.isoformat()} looks wrong — ignored"))
        return issues

    # --- resolution helpers ---
    @staticmethod
    def _match_niche(value: str, ctx: ResolveContext) -> Niche | None:
        for piece in value.split(","):
            niche = ctx.niches.get(piece.strip().lower())
            if niche is not None:
                return niche
        return None

    @staticmethod
    def _match_country(value: str, ctx: ResolveContext) -> Country | None:
        for piece in value.split(","):
            key = piece.strip().lower()
            key = COUNTRY_ALIASES.get(key, key)
            country = ctx.countries.get(key)
            if country is not None:
                return country
        return None

    @staticmethod
    def _number(value: Any) -> float | None:
        if value is None or value == "":
            return None
        if isinstance(value, int | float):
            return float(value)
        return parse_number(str(value))

    def apply(
        self, db: Session, company_id: uuid.UUID, user_id: uuid.UUID,
        canonical: dict[str, Any], ctx: ResolveContext,
    ) -> ApplyOutcome:
        name = (canonical.get("name") or "").strip()
        existing = ctx.projects_by_name.get(name.lower())
        snapshot = None
        if existing is not None:
            snapshot = {
                "name": existing.name,
                "main_niche_id": existing.main_niche_id,
                "project_niche_id": existing.project_niche_id,
                "target_country_id": existing.target_country_id,
                "assignee_id": str(existing.assignee_id) if existing.assignee_id else None,
                "team_lead_id": str(existing.team_lead_id) if existing.team_lead_id else None,
                "target_links": existing.target_links,
                "monthly_budget": float(existing.monthly_budget or 0),
                "goal": existing.goal,
                "status": existing.status,
                "notes": existing.notes,
            }
            project = existing
        else:
            project = Project(company_id=company_id, created_by=user_id, name=name)
            db.add(project)
            ctx.projects_by_name[name.lower()] = project

        main = self._match_niche((canonical.get("main_niche") or ""), ctx)
        if main is not None:
            project.main_niche_id = main.id
        proj_niche = self._match_niche((canonical.get("project_niche") or ""), ctx)
        if proj_niche is not None:
            project.project_niche_id = proj_niche.id
        country = self._match_country((canonical.get("target_country") or ""), ctx)
        if country is not None:
            project.target_country_id = country.id
        assignee = ctx.users.get((canonical.get("assignee") or "").strip().lower())
        if assignee is not None:
            project.assignee_id = assignee.id
        lead = ctx.users.get((canonical.get("team_lead") or "").strip().lower())
        if lead is not None:
            project.team_lead_id = lead.id

        try:
            target_links = self._number(canonical.get("target_links"))
        except ValueError:
            target_links = None
        # Guard against misaligned source cells (e.g. a date serial under "links").
        if target_links is not None and 0 <= target_links <= 100000:
            project.target_links = int(target_links)
        try:
            budget = self._number(canonical.get("monthly_budget"))
        except ValueError:
            budget = None
        if budget is not None and budget >= 0:
            project.monthly_budget = budget
        goal = canonical.get("goal")
        if goal not in (None, ""):
            project.goal = str(goal)
        due = canonical.get("due_date")
        if isinstance(due, date) and 2000 <= due.year <= 2100:
            project.due_date = due
        status = (canonical.get("status") or "").strip().lower()
        if status in PROJECT_STATUSES:
            project.status = status
        notes = canonical.get("notes")
        if notes not in (None, ""):
            project.notes = str(notes)

        db.flush()
        return ApplyOutcome(
            action="updated" if existing is not None else "created",
            entity_id=project.id,
            old_snapshot=snapshot,
        )


class ProjectsTemplateProfile(ProjectProfileBase):
    """The clean Projects template/export format (single sheet, canonical headers)."""

    key = "projects_template"
    label = "Projects — standard template"
    description = "The Projects template/export columns (name, niches, country, assignee, …)."
    column_mapping = {
        "name": "name", "main_niche": "main niche", "project_niche": "project niche",
        "target_country": "country", "assignee": "assignee email", "team_lead": "team lead email",
        "target_links": "target links", "monthly_budget": "monthly budget", "goal": "goal",
        "due_date": "due date", "status": "status", "notes": "notes",
    }

    def extract(self, filename: str, content: bytes) -> list[ExtractedRow]:
        rows = parse_table(filename, content)
        out: list[ExtractedRow] = []
        for i, row in enumerate(rows, start=2):
            if not (row.get("name") or "").strip():
                continue
            canonical = {k: row.get(k, "") for k in self.column_mapping}
            canonical["due_date"] = parse_date(row.get("due_date", "")) if row.get("due_date") else None
            out.append(ExtractedRow(row_number=i, raw=dict(row), canonical=canonical))
        return out


class MasterProjectsProfile(ProjectProfileBase):
    """The team's real Master workbook 'Projects & Assignee' tab.

    Maps the fixed leading columns to Projects and ignores the trailing per-month
    Goal/Actual columns. Handles Excel-serial due dates and multi-value cells.
    """

    key = "master_projects"
    label = "Master Sheet — Projects & Assignee"
    description = "Your Master workbook's 'Projects & Assignee' tab (skips the per-month columns)."
    column_mapping = {
        "name": "Projects", "main_niche": "Main Niche", "project_niche": "Project Niche",
        "target_country": "Target Country", "assignee": "Assignee", "team_lead": "Team Lead",
        "target_links": "Target Links", "monthly_budget": "Monthly Budget", "goal": "Goal",
        "due_date": "Due Date",
    }

    def extract(self, filename: str, content: bytes) -> list[ExtractedRow]:
        headers, data = _read_named_sheet(content, "Projects & Assignee", "Projects and Assignee")
        index = {h.strip().lower(): i for i, h in enumerate(headers) if h}

        def get(values: list[Any], header: str) -> Any:
            i = index.get(header.lower())
            if i is None or i >= len(values):
                return ""
            return values[i]

        out: list[ExtractedRow] = []
        for excel_row, values in data:
            name = get(values, "Projects")
            name = str(name).strip() if name is not None else ""
            if not name:
                continue  # blank / spacer row
            raw = {h: values[i] if i < len(values) else None for i, h in enumerate(headers) if h}
            canonical: dict[str, Any] = {
                "name": name,
                "main_niche": _s(get(values, "Main Niche")),
                "project_niche": _s(get(values, "Project Niche")),
                "target_country": _s(get(values, "Target Country")),
                "assignee": _s(get(values, "Assignee")),
                "team_lead": _s(get(values, "Team Lead")),
                "target_links": get(values, "Target Links"),
                "monthly_budget": get(values, "Monthly Budget"),
                "goal": _s(get(values, "Goal")),
                "due_date": _safe_date(get(values, "Due Date")),
                "status": "active",
                "notes": None,
            }
            out.append(
                ExtractedRow(row_number=excel_row, raw=_jsonable(raw), canonical=canonical)
            )
        return out


def _s(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _safe_date(value: Any) -> date | None:
    try:
        return _coerce_date(value)
    except (ValueError, OverflowError):
        return None


def _jsonable(raw: dict[str, Any]) -> dict[str, Any]:
    out: dict[str, Any] = {}
    for key, value in raw.items():
        if isinstance(value, datetime | date):
            out[key] = value.isoformat()
        else:
            out[key] = value
    return out


# ---------------------------------------------------------------------------
# Payments ledger profile — the team's 50-tab "Guest Post Payments" workbook,
# where each tab is one client and each row is one payment.
# ---------------------------------------------------------------------------

# Tabs that are not client ledgers.
LEDGER_SKIP_TABS = {"index", "from feb ali projects"}
# Header cells that identify the real header row within a tab.
_HEADER_HINTS = ("live link", "amount", "mode of payment")


@dataclass
class PaymentContext:
    clients: dict[str, Client] = field(default_factory=dict)
    users: dict[str, User] = field(default_factory=dict)
    seen: set[str] = field(default_factory=set)  # dedupe keys already in the DB


def _classify_status(remarks: str, mode: str, amount: float | None) -> str:
    r = (remarks or "").strip().lower()
    m = (mode or "").strip().lower()
    if "cancel" in r:
        return "cancelled"
    if "reject" in r:
        return "rejected"
    if "negotiat" in r or "in talk" in r or "discussion" in r:
        return "negotiation"
    if "free" in r or "free" in m or "no payment" in m or m in ("-", "--"):
        return "free"
    if "done" in r or "paid" in r:
        return "paid"
    if amount and amount > 0:
        return "paid"
    if not amount:
        return "free" if not r else "pending"
    return "pending"


class PaymentsLedgerProfile:
    """Reads every client tab of the Payments workbook into per-client payments."""

    key = "payments_ledger"
    label = "Payments — per-client ledger tabs"
    description = "Your Payments workbook: each tab is a client, each row a payment."
    entity_type = "payment"
    on_duplicate = "skip"  # re-importing the same workbook skips rows already loaded
    column_mapping = {
        "client": "(tab name)", "payment_date": "Date", "live_link": "Live Links",
        "mode_of_payment": "Mode of Payment", "amount": "Amount ($)",
        "amount_inr": "INR / CAD", "remarks": "Remarks",
        "transaction_id": "Transaction ID", "notified": "Notified?",
        "website": "Website URL", "member": "Member Name",
    }

    # --- extract ---
    def extract(self, filename: str, content: bytes) -> list[ExtractedRow]:
        try:
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        except Exception as exc:  # noqa: BLE001
            raise BadRequest("Could not read the Excel file") from exc
        out: list[ExtractedRow] = []
        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() in LEDGER_SKIP_TABS:
                continue
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            header_idx = self._find_header(rows)
            if header_idx is None:
                continue
            headers = [(_s(h)).lower() for h in rows[header_idx]]
            index = {h: i for i, h in enumerate(headers) if h}
            for offset, values in enumerate(rows[header_idx + 1 :], start=header_idx + 2):
                if values is None or all(c is None for c in values):
                    continue
                canonical = self._map_row(sheet_name, headers, index, list(values))
                if canonical is None:
                    continue
                raw = _jsonable(
                    {h: values[i] if i < len(values) else None for i, h in enumerate(headers) if h}
                )
                raw["__client__"] = sheet_name
                out.append(
                    ExtractedRow(row_number=offset, raw=raw, canonical=canonical, source=sheet_name)
                )
        wb.close()
        return out

    @staticmethod
    def _find_header(rows: list) -> int | None:
        for i, row in enumerate(rows[:8]):
            cells = [(_s(c)).lower() for c in (row or [])]
            if any(any(hint in c for hint in _HEADER_HINTS) for c in cells):
                return i
        return None

    def _map_row(self, client: str, headers, index, values) -> dict | None:
        def cell(*names: str) -> str:
            for name in names:
                for h, i in index.items():
                    if name in h:
                        return _s(values[i]) if i < len(values) else ""
            return ""

        live_link = cell("live link")
        amount = _safe_number(cell("amount ($)", "amount($)", "amount"))
        remarks = cell("remarks")
        mode = cell("mode of payment", "mode")
        # A row with no link and no amount is a spacer / note — skip it.
        if not live_link and amount is None and not remarks:
            return None
        local = _safe_number(cell("inr", "cad", "local"))
        return {
            "client": client.strip(),
            "live_link": live_link,
            "amount": amount,
            "amount_inr": local,
            "mode_of_payment": mode,
            "payment_date": _safe_date(self._raw_cell(index, values, "date")),
            "transaction_id": cell("transaction"),
            "remarks": remarks,
            "website": cell("website"),
            "member": cell("member"),
            "notified": _truthy(cell("notified")),
            "status": _classify_status(remarks, mode, amount),
        }

    @staticmethod
    def _raw_cell(index, values, name):
        for h, i in index.items():
            if name in h:
                return values[i] if i < len(values) else None
        return None

    # --- context / validate / dedupe / apply ---
    def build_context(self, db: Session, company_id: uuid.UUID) -> PaymentContext:
        ctx = PaymentContext()
        for c in db.query(Client).filter(Client.company_id == company_id).all():
            ctx.clients[c.name.strip().lower()] = c
        for u in db.query(User).filter(User.company_id == company_id).all():
            ctx.users[u.full_name.strip().lower()] = u
            ctx.users[u.email.lower()] = u
        for p in db.query(Payment).filter(Payment.company_id == company_id).all():
            ctx.seen.add(self._key(p.client.name if p.client else "", p.live_link, p.amount, p.payment_date))
        return ctx

    @staticmethod
    def _key(client, live_link, amount, pay_date) -> str:
        d = pay_date.isoformat() if hasattr(pay_date, "isoformat") else (pay_date or "")
        # Normalize amount to a fixed format so a stored Decimal and an extracted
        # float compare equal (e.g. Decimal('40.00') and 40.0 both -> "40.00").
        amt = "" if amount in (None, "") else f"{float(amount):.2f}"
        return f"{(client or '').strip().lower()}|{(live_link or '').strip().lower()}|{amt}|{d}"

    def dedupe_key(self, canonical: dict) -> str:
        return self._key(
            canonical.get("client"), canonical.get("live_link"),
            canonical.get("amount"), canonical.get("payment_date"),
        )

    def exists(self, key: str, ctx: PaymentContext) -> bool:
        return key in ctx.seen

    def validate(self, canonical: dict, ctx: PaymentContext) -> list[Issue]:
        issues: list[Issue] = []
        if not canonical.get("client"):
            issues.append(Issue("error", "Client (tab name) is missing"))
        if canonical.get("amount") is None and canonical.get("status") != "free":
            issues.append(Issue("warning", "No amount — treated as the classified status"))
        member = (canonical.get("member") or "").strip()
        if member and ctx.users.get(member.lower()) is None:
            issues.append(Issue("warning", f"Member '{member}' not matched to a user"))
        return issues

    def apply(self, db, company_id, user_id, canonical, ctx: PaymentContext) -> ApplyOutcome:
        client_name = (canonical.get("client") or "").strip()
        client = ctx.clients.get(client_name.lower())
        if client is None:
            client = Client(company_id=company_id, name=client_name, created_by=user_id)
            db.add(client)
            db.flush()
            ctx.clients[client_name.lower()] = client

        member = ctx.users.get((canonical.get("member") or "").strip().lower())
        amount = canonical.get("amount")
        amount_usd = round(float(amount), 2) if amount is not None else None
        status = canonical.get("status") if canonical.get("status") in PAYMENT_STATUSES else "pending"
        payment = Payment(
            company_id=company_id,
            client_id=client.id,
            created_by=user_id,
            currency="USD",
            amount=amount,
            fx_to_usd=1.0 if amount is not None else None,
            amount_usd=amount_usd,
            amount_inr=canonical.get("amount_inr"),
            live_link=_clip(canonical.get("live_link"), 700),
            mode_of_payment=_clip(canonical.get("mode_of_payment"), 255),
            payment_date=canonical.get("payment_date"),
            transaction_id=_clip(canonical.get("transaction_id"), 120),
            remarks=canonical.get("remarks") or None,
            notified=bool(canonical.get("notified")),
            attributed_to_id=member.id if member else None,
            status=status,
        )
        db.add(payment)
        db.flush()
        db.add(
            PaymentStatusHistory(
                payment_id=payment.id, from_status=None, to_status=status,
                changed_by=user_id, note="imported",
            )
        )
        ctx.seen.add(self.dedupe_key(canonical))
        return ApplyOutcome(action="created", entity_id=payment.id)


def _safe_number(value) -> float | None:
    try:
        return parse_number(str(value)) if value not in (None, "") else None
    except ValueError:
        return None


def _truthy(value: str) -> bool:
    return (value or "").strip().lower() in {"yes", "y", "true", "1", "done", "notified"}


def _clip(value: str | None, length: int) -> str | None:
    """Trim a value to fit its column; returns None for blanks."""
    if not value:
        return None
    text = str(value).strip()
    return text[:length] if text else None


# --- registry ---
_PROFILES: dict[str, object] = {
    p.key: p
    for p in (ProjectsTemplateProfile(), MasterProjectsProfile(), PaymentsLedgerProfile())
}


def get_profile(key: str):
    profile = _PROFILES.get(key)
    if profile is None:
        raise BadRequest(f"Unknown import profile '{key}'")
    return profile


def list_profiles() -> list:
    return list(_PROFILES.values())
